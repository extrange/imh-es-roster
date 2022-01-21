"""
Extract the schedule from the XLSX roster to a JSON file
"""
import json
from pathlib import Path

import typer
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print
from rich.table import Table

# The row in the sheet to extract dates from
DATE_ROW = 5


def find_name(ws: Worksheet, name: str) -> tuple[str, int]:
    """
    Find the row with the schedule for `name`

    Must be an exact match
    """
    for row in ws.rows:
        for cell in row:
            if str(cell.value).strip() == name:
                return (cell.column_letter, cell.row)
    else:
        raise Exception(f"Couldn't find '{name}' in {ws}")


def find_filled_cell(row, first=True) -> str | None:
    """
    If first==True, find first column in the row which has a value
    Else find the last
    """
    for cell in row if first else reversed(row):
        if cell.value:
            return cell.column_letter


def print_schedule(schedule) -> None:
    """
    Pretty-print the output JSON for user verification.
    """
    table = Table(header_style="bold magenta")
    table.add_column("Day")
    table.add_column("Shift")
    for day, shift in schedule:
        table.add_row(str(day), shift)
    print(table)


def main(
    file: Path = typer.Argument(..., help="Path to the Excel roster file"),
    name: str = typer.Option(
        ..., help="Name to extract schedule for. Note: Must be an exact match"
    ),
) -> None:
    """
    Extracts the schedule from the .xlsx roster to a JSON file.

    Doesn't care about what month it is.
    """
    wb = load_workbook(str(file))
    ws = wb[wb.sheetnames[0]]  # Get first sheet

    # Detect row to extract schedule from
    schedule_row = find_name(ws, name)[1]

    # Detect start and end column of date range
    date_start_col = find_filled_cell(ws[DATE_ROW])
    date_end_col = find_filled_cell(ws[DATE_ROW], False)

    schedule = []
    for date_cell in ws[f"{date_start_col}{DATE_ROW}":f"{date_end_col}{DATE_ROW}"][0]:
        current_col = date_cell.column_letter
        schedule_cell = ws[f"{current_col}{schedule_row}"]
        # Add only if there is a value in my schedule
        if schedule_cell.value:
            schedule.append((date_cell.value, schedule_cell.value))

    print_schedule(schedule)

    output_path = Path(file.stem + ".json")
    json.dump(schedule, output_path.open("w", encoding="utf8"))
    print(f"Wrote to [cyan]{output_path}[/cyan]")


if __name__ == "__main__":
    app = typer.Typer(add_completion=False)
    app.command()(main)
    app()
